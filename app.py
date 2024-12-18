from flask import Flask, render_template, request, redirect, jsonify, send_file
import sqlite3
import io
import openpyxl
from openpyxl import Workbook
import os

app = Flask(__name__)

app.config['UPLOAD_FOLDER'] = 'uploads'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Configuração inicial do banco de dados
def init_db():
    with sqlite3.connect('finance.db') as conn:
        cursor = conn.cursor()
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS transactions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT NOT NULL,
                quantia NUMERIC NOT NULL,
                description TEXT NOT NULL,
                value REAL NOT NULL,
                payment_method TEXT NOT NULL,
                type TEXT NOT NULL
            )
        ''')
        conn.commit()

init_db()

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        date = request.form['date']
        quantia = request.form['quantia']
        description = request.form['description']
        value = float(request.form['value'])
        payment_method = request.form['payment_method']
        type_ = request.form['type']

        with sqlite3.connect('finance.db') as conn:
            cursor = conn.cursor()
            cursor.execute('''
                INSERT INTO transactions (date, quantia, description, value, payment_method, type)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (date, quantia, description, value, payment_method, type_))
            conn.commit()

        return redirect('/')

    search_query = request.args.get('search', '')  # Captura a busca inteligente
    month_filter = request.args.get('month')  # Captura o filtro de mês (se houver)
    payment_filter = request.args.get('payment_method', '')  # Pega o filtro da URL
    tipe_filter = request.args.get('type', '')
    payment_methods = ["Pix", "Dinheiro", "Credito_c6", "Credito_nu"]
    tipos = ["Gasto", "Ganho"]

    query = "SELECT * FROM transactions WHERE 1=1"  # Consulta base
    params = []

    if payment_filter:
        query += " AND payment_method = ?"
        params.append(payment_filter)
        
    if month_filter:
        query += " AND strftime('%m', date) = ?"
        params.append(month_filter)

    if tipe_filter:
        query += " AND type = ?"
        params.append(tipe_filter)

    # Implementação da busca inteligente
    if search_query:
        if search_query.startswith('>'):
            try:
                value = float(search_query[1:])  # Captura o valor após o `>`
                query += ' AND value > ?'
                params.append(value)
            except ValueError:
                pass  # Ignorar se não for um número válido
        elif search_query.startswith('<'):
            try:
                value = float(search_query[1:])  # Captura o valor após o `<`
                query += ' AND value < ?'
                params.append(value)
            except ValueError:
                pass  # Ignorar se não for um número válido
        else:
            # Busca em outros campos: descrição, data, método de pagamento
            query += '''
                AND (description LIKE ? OR 
                     date LIKE ? OR 
                     payment_method LIKE ? OR 
                     CAST(value AS TEXT) LIKE ?)
            '''
            search_term = f'%{search_query}%'  # Termo genérico
            params.extend([search_term, search_term, search_term, search_term])

    query += ' ORDER BY date'

    with sqlite3.connect('finance.db') as conn:
        cursor = conn.cursor()
        cursor.execute(query, params)
        transactions = cursor.fetchall()

        # Obter todos os meses únicos disponíveis
        cursor.execute("SELECT DISTINCT strftime('%m', date) as month FROM transactions ORDER BY month")
        months = [row[0] for row in cursor.fetchall()]

    return render_template('index.html', transactions=transactions, months=months, month_filter=month_filter, payment_filter=payment_filter, payment_methods=payment_methods, tipos=tipos, tipe_filter=tipe_filter)

@app.route('/generate_excel', methods=['GET'])
def generate_excel():
    # Conectar ao banco de dados e obter as transações
    with sqlite3.connect('finance.db') as conn:
        cursor = conn.cursor()
        cursor.execute("SELECT id, date, quantia, description, value, payment_method, type FROM transactions")
        transactions = cursor.fetchall()

    # Criar um arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Transações"

    # Cabeçalho da planilha
    ws.append(["ID", "Data", "Quantia", "Descrição", "Valor", "Método de Pagamento", "Tipo"])

    # Adicionar os dados das transações
    for transaction in transactions:
        ws.append(transaction)

    # Salvar a planilha em memória (em vez de salvar fisicamente)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar a planilha para o cliente
    return send_file(output, as_attachment=True, download_name="transacoes.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

def save_excel_to_db(file_path):
    # Carregar a planilha
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    # Iterar pelas linhas da planilha
    with sqlite3.connect('finance.db') as conn:
        cursor = conn.cursor()
        for row in ws.iter_rows(min_row=2, values_only=True):  # Ignorar cabeçalho
            cursor.execute('''
                INSERT INTO transactions (id, date, quantia, description, value, payment_method, type)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', row)
        conn.commit()

# Rota para upload da planilha
@app.route('/upload_excel', methods=['POST'])
def upload_excel():
    file = request.files['file']
    if file and file.filename.endswith('.xlsx'):
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(file_path)
        
        # Salvar os dados da planilha no banco
        save_excel_to_db(file_path)
        
        # Redirecionar para a página inicial após o upload
        return redirect('/')

    return "Arquivo inválido. Por favor, envie um arquivo .xlsx."

# Rota para baixar a planilha de exemplo
@app.route('/download_example')
def download_example():
    # Criar um arquivo de exemplo
    wb = Workbook()
    ws = wb.active
    ws.title = "Transações Exemplo"
    
    # Cabeçalho da planilha
    ws.append(["ID", "Data", "Quantia", "Descrição", "Valor", "Método de Pagamento", "Tipo"])
    
    # Dados de exemplo
    ws.append([1, "2023-01-01", 100, "Exemplo de Descrição", 50.5, "Pix", "Gasto"])
    
    # Salvar o arquivo em memória
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(output, as_attachment=True, download_name="exemplo_transacoes.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.route('/delete/<int:id>', methods=['POST'])
def delete_transaction(id):
    with sqlite3.connect('finance.db') as conn:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM transactions WHERE id = ?', (id,))
        conn.commit()
    return jsonify({'success': True})

@app.route('/get_transaction/<int:id>', methods=['GET'])
def get_transaction(id):
    with sqlite3.connect('finance.db') as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM transactions WHERE id = ?', (id,))
        transaction = cursor.fetchone()
    return jsonify(transaction)

@app.route('/edit/<int:id>', methods=['POST'])
def edit_transaction(id):
    data = request.json
    with sqlite3.connect('finance.db') as conn: 
        cursor = conn.cursor()
        cursor.execute('''
            UPDATE transactions
            SET date = ?, quantia =?, description = ?, value = ?, payment_method = ?, type = ?
            WHERE id = ?
        ''', (data['date'], data['quantia'], data['description'], data['value'], data['payment_method'], data['type'], id))
        conn.commit()
    return jsonify({'success': True})

if __name__ == '__main__':
    app.run(debug=True)
